# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, Http404, HttpResponse
from django.shortcuts import render
from orders.models import Order, OrderDetail
from bills.models import Bill, BillDetail
from requests.models import Request, RequestDetail
from projects.models import Project
from suppliers.models import Supplier
from products.models import Product
from activities.models import Activity
from datetime import date
import datetime
from operator import itemgetter


@login_required
def report_create(request):

    return render(request, '../templates/reports/template.jade')


def ordersbyproject(request, project):

    orders = Order.objects.filter(order_project=project)

    subtotal = 0
    iv = 0
    total = 0

    for order in orders:
        subtotal = subtotal+order.order_subtotal
        iv = iv+order.order_iv
        total = total+order.order_total

    # details = orders.order_product_list.all()
    return render(request, '../templates/reports/ordersbyproject.jade', {'orders': orders, 'subtotal':subtotal, 'iv':iv,
                                                                   'total':total})

def billsbyproject(request, project):

    bills = Bill.objects.filter(bill_order__order_project=project)

    subtotal = 0
    iv = 0
    total = 0

    for bill in bills:
        subtotal = subtotal + bill.bill_subtotal
        iv = iv + bill.bill_iv
        total = total + bill.bill_total

    # details = orders.order_product_list.all()
    return render(request, '../templates/reports/billsbyproject.jade', {'bills': bills, 'subtotal': subtotal, 'iv': iv,
                                                               'total': total})

def byorder(request, order):


    order = Order.objects.get(pk=order)

    bills = Bill.objects.filter(bill_order=order)

    orderdetail = order.order_product_list.all().order_by('order_detail_product_code')

    bill_details = []

    ivbill = 0
    subtotalbill=0
    totalbill = 0

    for bill in bills:

        billdetails = bill.bill_detail_list.all()
        totalbill = totalbill+bill.bill_total
        ivbill = ivbill + bill.bill_iv
        subtotalbill = subtotalbill + bill.bill_subtotal


        for billdetail in billdetails:

            bill_details.append(billdetail)

    difference = totalbill - order.order_total


    return render(request, '../templates/reports/byordersimplified.jade', {'order':order, 'bills': bills,
                                                                           'detail': orderdetail, 'bdetails':bill_details,
                                                                           'ivbill':ivbill, 'subtotalbill':subtotalbill,
                                                                       'totalbill':totalbill, 'difference':difference})

def generate_xlsx(report_name, objects, project, activity, supplier, date, date_ini, date_end ):

    wb = Workbook()
    current_row = 1
    current_col = 1
    wb._active_sheet_index = 0
    #get the first sheet, created per default
    ws = wb.active
    #change the title of the sheet
    ws.title = 'Reporte'

    ws.cell(row=current_row, column=1, value='COTO COMPANY')

    current_row += 2

    ws.cell(row=current_row, column=1, value='REPORTE:')
    ws.cell(row=current_row, column=2, value=report_name)
    current_row += 1

    ws.cell(row=current_row, column=1, value='FECHA:')
    ws.cell(row=current_row, column=2, value=date)
    current_row += 1

    ws.cell(row=current_row, column=1, value='PROYECTO:')
    ws.cell(row=current_row, column=2, value=project)
    current_row += 1

    ws.cell(row=current_row, column=1, value='ACTIVIDAD:')
    ws.cell(row=current_row, column=2, value=activity)
    current_row += 1

    ws.cell(row=current_row, column=1, value='PROVEEDOR:')
    ws.cell(row=current_row, column=2, value=supplier)
    current_row += 1

    ws.cell(row=current_row, column=1, value='FECHA INICIAL:')
    ws.cell(row=current_row, column=2, value=date_ini)
    current_row += 1

    ws.cell(row=current_row, column=1, value='FECHA FINAL:')
    ws.cell(row=current_row, column=2, value=date_end)
    current_row += 2

    # Header row

    ws.cell(row=current_row, column=1, value='Código')
    ws.cell(row=current_row, column=2, value='Descripción')
    ws.cell(row=current_row, column=3, value='Unidad')
    ws.cell(row=current_row, column=4, value='Presupuestado')
    ws.cell(row=current_row, column=5, value='Ordenado')
    ws.cell(row=current_row, column=6, value='Facturado')
    ws.cell(row=current_row, column=7, value='En Tránsito')
    ws.cell(row=current_row, column=8, value='Requisado')
    ws.cell(row=current_row, column=9, value='Inventario')

    current_row += 1

    for item in objects:
        ws.cell(row=current_row, column=1, value=item['code'])
        ws.cell(row=current_row, column=2, value=item['description'])
        ws.cell(row=current_row, column=3, value=item['unit'])
        ws.cell(row=current_row, column=4, value='-')
        ws.cell(row=current_row, column=5, value=item['ordered'])
        ws.cell(row=current_row, column=6, value=item['billed'])
        ws.cell(row=current_row, column=7, value=item['ordered'] - item['billed'])
        ws.cell(row=current_row, column=8, value='-')
        ws.cell(row=current_row, column=9, value='-')
        current_row += 1

   

    return wb


def generalreport(request):

    # if request.is_ajax():

    type = request.GET['type']
    project = request.GET['project']
    activity = request.GET['activity']
    supplier = request.GET['supplier']
    product = request.GET['product']
    date_ini = request.GET['date_ini']
    date_end = request.GET['date_end']
    to_excel = request.GET['to_excel']

    today = date.today()

    projectd = "Todos(as)"
    activityd = "Todos(as)"
    supplierd = "Todos(as)"
    productd = "Todos(as)"

    bills = Bill.objects.all()

    if project != '0':
        bills = bills.filter(bill_order__order_project=project)
        projectobj = Project.objects.get(id=project)
        projectd = '%s - %s' % (projectobj.id, projectobj.project_name)

    if activity != '0':
        bills = bills.filter(bill_order__order_activity=activity)
        activityobj = Activity.objects.get(id=activity)
        activityd = '%s - %s' % (activityobj.id, activityobj.activity_name)

    if supplier != '0':
        bills = bills.filter(bill_supplier=supplier)
        supplierobj = Supplier.objects.get(id=supplier)
        supplierd = '%s ' % supplierobj.supplier_name

    orders = Order.objects.all()

    requests = Request.objects.all()


    if project != '0':
        orders = orders.filter(order_project=project)
        projectobj = Project.objects.get(id=project)
        projectd = '%s - %s' % (projectobj.id, projectobj.project_name)

    if activity != '0':
        orders = orders.filter(order_activity=activity)
        activityobj = Activity.objects.get(id=activity)
        activityd = '%s - %s' % (activityobj.id, activityobj.activity_name)

    if supplier != '0':
        orders = orders.filter(order_supplier=supplier)
        supplierobj = Supplier.objects.get(id=supplier)
        supplierd = '%s ' % supplierobj.supplier_name


    if date_ini and date_end:

        orders=orders.filter(order_date__range=[date_ini, date_end])
        bills = bills.filter(bill_date__range=[date_ini, date_end])
        requests = requests.filter(request_date__range=[date_ini, date_end])

        date_ini = datetime.datetime.strptime(date_ini, "%Y-%m-%d").date()
        date_end = datetime.datetime.strptime(date_end, "%Y-%m-%d").date()

    else:
        date_ini = '-'
        date_end = '-'


    if type == '1':


        ivbill = 0
        totalbill = 0

        for bill in bills:

            totalbill = totalbill + bill.bill_total
            ivbill = ivbill + bill.bill_iv

        return render(request, '../templates/reports/general_bills.jade', {'bills': bills,'total':totalbill,
                                                                            'iv':ivbill,'project':projectd,
                                                                            'activity':activityd, 'supplier':supplierd,
                                                                            'date':today, 'date_ini':date_ini,
                                                                            'date_end':date_end })

    if type == '2':

        ivbill = 0
        totalbill = 0
        productsbill = 0

        for bill in bills:

            billdetails = bill.bill_detail_list.all()

            if product == '0':
                ivbill = ivbill + bill.bill_iv

            for billdetail in billdetails:

                if product == '0':

                    totalbill = totalbill + billdetail.bill_detail_total
                    productsbill = productsbill + billdetail.bill_detail_amount

                else:
                    if billdetail.bill_detail_product_code == product:

                        productobj = Product.objects.get(product_code=product)
                        productd = '%s - %s' % (product, productobj.product_description)

                        totalbill = totalbill + billdetail.bill_detail_total
                        ivbill = ivbill + ((billdetail.bill_detail_iv/100)*billdetail.bill_detail_total)
                        productsbill = productsbill + billdetail.bill_detail_amount

        return render(request, '../templates/reports/detailed_bills.jade', {'bills': bills, 'subtotal':totalbill,
                                                                            'total': totalbill+ivbill,
                                                                            'iv':ivbill, 'cantprod':productsbill,
                                                                            'product':product, 'project':projectd,
                                                                            'activity':activityd, 'supplier':supplierd,
                                                                            'productd':productd, 'date':today,
                                                                            'date_ini':date_ini, 'date_end':date_end})

    if type == '3':

        ivorder = 0
        totalorder = 0

        for order in orders:
            totalorder = totalorder + order.order_total
            ivorder = ivorder + order.order_iv

        return render(request, '../templates/reports/general_orders.jade', {'orders': orders, 'total': totalorder,
                                                                            'iv': ivorder, 'project': projectd,
                                                                            'activity': activityd,
                                                                            'supplier': supplierd,
                                                                            'date': today,'date_ini':date_ini,
                                                                            'date_end':date_end})

    if type == '4':

        ivorder = 0
        totalorder = 0
        productsorder = 0


        for order in orders:

            orderdetails = order.order_product_list.all()

            if product == '0':
                ivorder = ivorder + order.order_iv

            for orderdetail in orderdetails:

                if product == '0':

                    totalorder = totalorder + orderdetail.order_detail_total
                    ivorder = ivorder + orderdetail.order_detail_iv
                    productsorder = productsorder + orderdetail.order_detail_amount

                else:
                    if orderdetail.order_detail_product_code == product:

                        productobj = Product.objects.get(product_code=product)
                        productd = '%s - %s' % (product, productobj.product_description)

                        totalorder = totalorder + orderdetail.order_detail_total
                        ivorder = ivorder + ((orderdetail.order_detail_iv / 100) * orderdetail.order_detail_total)
                        productsorder = productsorder + orderdetail.order_detail_amount

        return render(request, '../templates/reports/detailed_orders.jade', {'orders': orders, 'subtotal':totalorder,
                                                                            'total': totalorder+ivorder,
                                                                            'iv':ivorder, 'cantprod':productsorder,
                                                                            'product':product, 'project':projectd,
                                                                            'activity':activityd, 'supplier':supplierd,
                                                                            'productd':productd, 'date':today,
                                                                            'date_ini':date_ini, 'date_end':date_end})

    if type == '5':

        details_array=[]

        for order in orders:

            order_detail = order.order_product_list.all()

            for detail in order_detail:

                i = 0

                for list_ in details_array:
                    if detail.order_detail_product_code in list_:
                        break
                    i += 1

                if len(details_array) == i:

                    details_array.append([detail.order_detail_product_code, detail.order_detail_description, 0,detail.order_detail_amount, 0,0,0,0,[order.id],[],[]])

                else:
                    details_array[i][3]=details_array[i][3]+detail.order_detail_amount
                    details_array[i][8].append(order.id)

        for bill in bills:

            bill_detail = bill.bill_detail_list.all()

            for detail in bill_detail:

                i = 0

                for list_ in details_array:
                    if detail.bill_detail_product_code in list_:
                        break
                    i += 1

                if len(details_array) == i:

                    details_array.append([detail.bill_detail_product_code, detail.bill_detail_description, 0,0, detail.bill_detail_amount,0,0,0,[],[bill.id],[]])

                else:
                    details_array[i][4]=details_array[i][4]+detail.bill_detail_amount
                    details_array[i][9].append(bill.id)

        for request2 in requests:

            request_detail = request2.request_product_list.all()

            for detail in request_detail:

                i = 0

                for list_ in details_array:
                    if detail.request_detail_product_code in list_:
                        break
                    i += 1

                if len(details_array) == i:

                    details_array.append([detail.request_detail_product_code, detail.request_detail_description, detail.request_detail_amount,0, 0,0,0,0,[],[],[request2.id]])

                else:
                    details_array[i][2]=details_array[i][2]+detail.request_detail_amount
                    details_array[i][10].append(request2.id)


        details_array = sorted(details_array, key=itemgetter(0))

        return render(request, '../templates/reports/general_byproduct.jade', {'array': details_array,
                                                                            'product': product, 'project': projectd,
                                                                            'activity': activityd,
                                                                            'supplier': supplierd,
                                                                            'productd': productd, 'date': today,
                                                                            'date_ini':date_ini, 'date_end':date_end})

    if type == '6':

        bills = bills.filter(bill_payed=False)

        totalbill = 0

        for bill in bills:

            if not bill.bill_debt:

                totalbill = totalbill + bill.bill_total

            else:

                totalbill = totalbill + bill.bill_debt

        print totalbill

        return render(request, '../templates/reports/general_debts.jade', {'bills': bills,
                                                                            'project': projectd,
                                                                            'activity': activityd,
                                                                            'supplier': supplierd,
                                                                            'date': today, 'total': totalbill})

    if type == '7':

        bills = bills.filter(bill_payed=True) | bills.filter(bill_half_payed=True)

        totalbill = 0

        for bill in bills:

            if bill.bill_payed:

                totalbill = totalbill + bill.bill_total

            else:

                totalbill = totalbill + bill.bill_total - bill.bill_debt

        print totalbill

        return render(request, '../templates/reports/general_payed.jade', {'bills': bills,
                                                                            'project': projectd,
                                                                            'activity': activityd,
                                                                            'supplier': supplierd,
                                                                            'date': today,
                                                                            'date_ini': date_ini,
                                                                            'date_end': date_end,
                                                                            'total': totalbill})

    if type == '8':

        arr = []

        for order in orders:

            order_detail = order.order_product_list.all()

            for detail in order_detail:

                i = 0

                if len(arr) > 0:
                    for element in arr:
                        if detail.order_detail_product_code == element['code']:
                            break
                        i += 1

                if len(arr) == i:

                    arr.append({'code': detail.order_detail_product_code,
                                'description': detail.order_detail_description,
                                'unit': detail.order_detail_unit,
                                'ordered': detail.order_detail_amount,
                                'billed': 0})

                else:
                    arr[i]['ordered'] = arr[i]['ordered'] + detail.order_detail_amount

        for bill in bills:

            bill_detail = bill.bill_detail_list.all()

            for detail in bill_detail:

                i = 0

                if len(arr) > 0:

                    for element in arr:
                        if detail.bill_detail_product_code == element['code']:
                            break
                        i += 1

                if len(arr) == i:

                    arr.append({'code': detail.bill_detail_product_code,
                                'description': detail.bill_detail_description,
                                'unit': detail.bill_detail_unit,
                                'ordered': 0,
                                'billed': detail.bill_detail_amount})

                else:
                    arr[i]['billed'] = arr[i]['billed'] + detail.bill_detail_amount

        arr = sorted(arr, key=itemgetter('description'))

        if to_excel == 'true':
            # now = datetime.datetime.now()
            
            # report = generate_xlsx('REPORTE DE RESUMEN DE MATERIALES', arr, projectd, activityd, supplierd, today, date_ini, date_end)
            
            # file_name = "Reporte Resumen de Materiales {}-{}-{}".format(now.year, now.month, now.day)

            # response = HttpResponse(content_type='application/ms-excel')
            # response['Content-Disposition'] = 'attachement; filename="{}"'.format(file_name)

            # report.save(response)
            # return response

            with NamedTemporaryFile() as temporary_file:
                now = datetime.datetime.now()
                report = generate_xlsx('REPORTE DE RESUMEN DE MATERIALES', arr, projectd, activityd, supplierd, today, date_ini, date_end)
                file_name = "Reporte Resumen de Materiales {}-{}-{}".format(now.year, now.month, now.day)
                report.save(temporary_file.name)
                temporary_file.seek(0)
                response = HttpResponse(temporary_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(file_name)
                return response

        return render(request, '../templates/reports/adrian.jade', {'objects': arr,
                                                                    'project': projectd,
                                                                    'activity': activityd,
                                                                    'supplier': supplierd,
                                                                    'date': today,
                                                                    'date_ini': date_ini,
                                                                    'date_end': date_end})

    if type == '9':

        ivbill = 0
        totalbill = 0

        activities9 = Activity.objects.all()
        bills9 = bills

        billsByActivity = []

        for bill in bills:

            totalbill = totalbill + bill.bill_total
            ivbill = ivbill + bill.bill_iv

        for activity9 in activities9:

            bills92 = bills9.filter(bill_order__order_activity=activity9)

            ivbill9 = 0
            totalbill9 = 0

            for bill in bills92:

                totalbill9 = totalbill9 + bill.bill_total
                ivbill9 = ivbill9 + bill.bill_iv

            if bills92:
                billsByActivity.append({'bills': bills92, 'total': totalbill9, 'totalIv': ivbill9,
                                        'activity': activity9})

        print(billsByActivity)

        return render(request, '../templates/reports/general_bills_category.jade', {'bills': billsByActivity,
                                                                                    'total': totalbill,
                                                                                    'iv': ivbill,
                                                                                    'project': projectd,
                                                                                    'activity': activityd,
                                                                                    'supplier': supplierd,
                                                                                    'date': today,
                                                                                    'date_ini': date_ini,
                                                                                    'date_end': date_end})

    if type == '10':

        ivbill = 0
        totalbill = 0
        
        activities10 = Activity.objects.all()
        bills10 = bills

        billsByActivity = []

        for bill in bills:

            totalbill = totalbill + bill.bill_total
            ivbill = ivbill + bill.bill_iv

        for activity10 in activities10:

            bills102 = bills10.filter(bill_order__order_activity=activity10)

            ivbill10 = 0
            totalbill10 = 0

            for bill in bills102:

                totalbill10 = totalbill10 + bill.bill_total
                ivbill10 = ivbill10 + bill.bill_iv

            if bills102:
                billsByActivity.append({'bills': bills102, 'total': totalbill10, 'totalIv': ivbill10,
                                        'activity': activity10})

        print(billsByActivity)

        return render(request, '../templates/reports/general_activities.jade', {'bills': billsByActivity,
                                                                                'total': totalbill,
                                                                                'iv': ivbill,
                                                                                'project': projectd,
                                                                                'activity': activityd,
                                                                                'supplier': supplierd,
                                                                                'date': today,
                                                                                'date_ini': date_ini,
                                                                                'date_end': date_end})

    # else:
    #     return False
